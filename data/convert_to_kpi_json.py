# data/convert_to_kpi_json.py
import json, os, math
from datetime import datetime, date
from typing import List, Tuple
import pandas as pd
from openpyxl import load_workbook

SOURCE_XLSX = "../NEW FORM Accident and Incident Investigations Action Tracker_ST0335-25x26sx (1).xlsx"
OUT_JSON    = "kpi_data.json"   # written into ./data

TODAY = date.today()

# ---- helper ---------------------------------------------------------------
def find_header_row(ws, expected_headers: List[str], max_search=300) -> Tuple[int, list]:
    """Find the first row that matches at least half of the expected headers."""
    for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_search, values_only=True), start=1):
        cells = [str(c).strip() if c is not None else "" for c in row]
        hits = sum(1 for h in expected_headers if h in cells)
        if hits >= max(2, math.ceil(0.5 * len(expected_headers))):
            return r_idx, cells
    raise RuntimeError("Could not find header row in sheet.")

def header_map(cells: list) -> dict:
    """Build col index -> header name map (1-based col index)."""
    mapping = {}
    for idx, name in enumerate(cells, start=1):
        if not name:
            continue
        s = str(name).strip()
        if s and not s.startswith("Unnamed"):
            mapping[idx] = s
    return mapping

def stream_table(ws, start_row: int, col_map: dict, key_cols: List[str], max_empty_seq=75, hard_cap=20000):
    """Stream rows from a sheet after header row until we encounter many empty rows."""
    out = []
    empty_seq = 0
    for r in ws.iter_rows(min_row=start_row+1, values_only=True):
        record = {}
        row_empty = True
        for c_idx, header in col_map.items():
            val = r[c_idx-1] if (c_idx-1) < len(r) else None
            # normalise excel datetime to date
            if isinstance(val, datetime):
                val = val.date()
            if val not in (None, ""):
                row_empty = False
            record[header] = val
        # skip if all key cols empty
        keys_empty = all((record.get(k) in (None, "")) for k in key_cols)
        if row_empty or keys_empty:
            empty_seq += 1
            if empty_seq >= max_empty_seq:
                break
            continue
        empty_seq = 0
        out.append(record)
        if len(out) >= hard_cap:
            break
    return out

# ---- main export -----------------------------------------------------------
def main():
    # open workbook in streaming mode
    wb = load_workbook(SOURCE_XLSX, data_only=True, read_only=True)

    ws_sa  = wb["Safety Action Tracking"]
    ws_inc = wb["Accident & Incident Detail"]

    sa_expected  = ["Status","Action/Recommendation Type","Date Action Raised","Period Action Raised",
                    "Dept","Function","Action Completion Target Date","Action Completed"]
    inc_expected = ["Status","Date","Reporting Period","Accident/Incident Type","Incident Type","Location",
                    "Person Type","Total Number of days Lost","RIDDOR","Investigation Due","Investigation Completion date","Function"]

    # detect headers
    sa_header_row, sa_header_cells   = find_header_row(ws_sa,  sa_expected)
    inc_header_row, inc_header_cells = find_header_row(ws_inc, inc_expected)

    sa_map  = header_map(sa_header_cells)
    inc_map = header_map(inc_header_cells)

    # key columns to detect ends
    sa_key_cols  = [c for c in ["Status","Action/Recommendation Type","Date Action Raised"] if c in sa_map.values()]
    inc_key_cols = [c for c in ["Status","Date","Incident Type","Accident/Incident Type"] if c in inc_map.values()]

    sa_rows  = stream_table(ws_sa,  sa_header_row,  sa_map,  sa_key_cols)
    inc_rows = stream_table(ws_inc, inc_header_row, inc_map, inc_key_cols)

    # keep only dashboard fields (rename some common variants)
    def pick(d, keys):
        return {k: d.get(k) for k in keys if k in d}

    sa_keep = [
        "Status","Action/Recommendation Type","Date Action Raised","Period Action Raised",
        "Investigation Ref (if applicable)","Incident Title","Recommendation Number",
        "Action/Recommendation","Dept","Action Owner","Function",
        "Action Completion Target Date","Action Completed","Comments",
        "Number of days taken to close","DWPI Days"
    ]
    inc_keep = [
        "Status","Date","Reporting Period",
        "Accident/Incident Reference","Accident/Incident Type","Incident Type","Incident",
        "Location","Category","Person Type","Description of incident",
        "Total Number of days Lost","RIDDOR","Function",
        "Investigating Manager","Investigation type","Investigation Due","Investigation Completion date",
        "Immediate Cause","Underlying Cause","Recommendations Made?","Comments","Claim received?"
    ]

    sa_out = []
    for r in sa_rows:
        row = pick(r, sa_keep)
        # derived overdue (calc)
        status = str(row.get("Status","")).strip().lower()
        tdate  = row.get("Action Completion Target Date")
        cdate  = row.get("Action Completed")
        is_overdue = (status != "closed") and isinstance(tdate, date) and (tdate < TODAY) and (cdate in (None, "") or (isinstance(cdate, date) and cdate > tdate))
        row["Is Overdue (calc)"] = is_overdue
        sa_out.append(row)

    inc_out = [pick(r, inc_keep) for r in inc_rows]

    payload = {
        "meta": {
            "source_file": os.path.basename(SOURCE_XLSX),
            "generated_utc": datetime.utcnow().isoformat() + "Z",
            "today": str(TODAY),
            "sheets": wb.sheetnames,
            "rows": {"safety_actions": len(sa_out), "incidents": len(inc_out)}
        },
        "safety_actions": sa_out,
        "incidents": inc_out
    }

    os.makedirs(os.path.dirname(OUT_JSON) or ".", exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, default=str)

    print(f"Wrote {OUT_JSON} with {len(sa_out)} actions and {len(inc_out)} incidents")

if __name__ == "__main__":
    main()
